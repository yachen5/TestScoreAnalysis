import datetime

import pandas as pd
import streamlit as st
from PIL import Image
from openpyxl import load_workbook

from LocalApps import SharedObjects


def display_image():
    """Displays an instructional image for file upload format."""
    image = Image.open("2023-12-23 013246.png")
    st.image(image, caption='上傳檔案格式如圖', width=800)


def get_test_date():
    """Prompts the user to input the test date."""
    col1, _ = st.columns(2)
    test_date = col1.date_input("本次考試日期為?", datetime.datetime.today())
    return test_date.isoformat()


def upload_excel_file():
    """Handles the Excel file upload and returns the file path."""
    return st.file_uploader("選擇一個 Excel 文件", type=["xlsx", "xls"])


def process_excel_file(excel_file_path, test_date):
    """
    Processes the uploaded Excel file and prepares data for analysis.

    Args:
        excel_file_path: Path to the uploaded Excel file.
        test_date: The test date provided by the user.

    Returns:
        A dictionary of processed data objects for subjects, years, and classes.
    """
    wb = load_workbook(excel_file_path, read_only=True)
    worksheet_names = wb.sheetnames
    li = []

    for sheet_name in worksheet_names:
        df_t = pd.read_excel(excel_file_path, sheet_name=sheet_name, header=None, engine='openpyxl')
        df_t = df_t.dropna(axis=1)
        df_t.columns = df_t.iloc[0]
        df_t = df_t[1:]
        df_t['年級'] = df_t['班級'].astype(str).str[0]
        df_t['學號'] = 'S' + df_t['學號'].astype(str)
        li.append(df_t)

    wb.close()
    df = pd.concat(li, ignore_index=True)
    df['年級_科目'] = df['年級'] + '_' + df['科目代號']
    df['考試日期'] = test_date
    df['年級'] = df['年級'].astype(int)
    df['班級'] = df['班級'].astype(str)
    df['座號'] = df['座號'].astype(str)

    # Split the '答題對錯' column into individual characters, each representing a question's correctness,
    # and create a new DataFrame where each column corresponds to a question.
    df_split = df['答題對錯'].apply(lambda x: pd.Series(list(x)))

    # Rename the columns of the new DataFrame to 'Q_01', 'Q_02', ..., based on the number of questions.
    df_split.columns = [f'Q_{str(n).zfill(2)}' for n in range(1, len(df_split.columns) + 1)]

    # Concatenate the original DataFrame with the new DataFrame and drop the original '答題對錯' column.
    df = pd.concat([df, df_split], axis=1).drop(columns=['答題對錯'])

    value_vars = [col for col in df.columns if col.startswith('Q')]
    melted_df = pd.melt(df, id_vars=[col for col in df.columns if col not in value_vars],
                        value_vars=value_vars, var_name='Question', value_name='Answer')
    melted_df = melted_df.drop('答題狀況', axis=1)

    result_dict = {}
    unique_values = melted_df['年級_科目'].unique()
    for value in unique_values:
        filtered_df = melted_df[melted_df['年級_科目'] == value].copy()
        filtered_df = filtered_df.dropna(subset='Answer')
        result_dict[value] = SharedObjects.Subject(filtered_df, value)

    year_dict = {}
    for a_year in list(df['年級'].unique()):
        year_obj = SharedObjects.ClassGroup(melted_df[melted_df['年級'] == a_year].copy(), a_year)
        year_obj.add_pw_by_subject(result_dict)
        year_dict[a_year] = year_obj

    class_dict = {}
    for a_class in list(melted_df['班級'].unique()):
        class_dict[a_class] = SharedObjects.Class(melted_df[melted_df['班級'] == a_class].copy())

    return {
        'subjects': result_dict,
        'year_groups': year_dict,
        'class_groups': class_dict
    }


def main():
    """Main function to handle the Streamlit app workflow."""
    st.title("使用 Streamlit 上傳和處理 Excel 文件")
    display_image()
    test_date = get_test_date()
    excel_file_path = upload_excel_file()

    if excel_file_path is not None:
        with st.spinner('Processing Data...'):
            data = process_excel_file(excel_file_path, test_date)
            st.session_state.update(data)
            st.success('Excel 檔案已經暫存到記憶體，準備進行下一步分析')


if __name__ == '__main__':
    main()
