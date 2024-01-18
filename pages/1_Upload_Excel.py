import datetime

import pandas as pd
import streamlit as st
from PIL import Image
from openpyxl import load_workbook

from LocalApps import AirtableTools as at
from LocalApps import SharedObjects


def main():
    st.title("使用 Streamlit 上傳和處理 Excel 文件")

    # Load your PNG image
    image = Image.open("2023-12-23 013246.png")

    # Display the image using st.image

    st.image(image, caption='上傳檔案格式如圖', width=800)
    col1, col2 = st.columns(2)
    test_date = col1.date_input("本次考試日期為?", datetime.datetime.today())
    test_date = test_date.isoformat()

    # Disable this function
    db_check = False
    # db_check = col2.checkbox('Load to DB', value=False)

    # File upload
    excel_file_path = st.file_uploader("選擇一個 Excel 文件", type=["xlsx", "xls"])

    if excel_file_path is not None:
        # Load the Excel workbook
        wb = load_workbook(excel_file_path, read_only=True)

        # Get the names of all worksheets in the Excel file
        worksheet_names = wb.sheetnames

        # Initialize an empty DataFrame to store the combined data
        li = []

        # Loop through each worksheet and append its data to the combined DataFrame
        for sheet_name in worksheet_names:
            # Read data from the current worksheet
            df_t = pd.read_excel(excel_file_path, sheet_name=sheet_name, header=None, engine='openpyxl')
            df_t = df_t.dropna(axis=1)

            # Use the first row as header
            df_t.columns = df_t.iloc[0]

            # Drop the first row (since it's now the header)
            df_t = df_t[1:]
            df_t['年級'] = df_t['班級'].astype(str).str[0]
            df_t['學號'] = 'S' + df_t['學號'].astype(str)

            # Append the data to the combined DataFrame
            li.append(df_t)

        # Close the workbook
        wb.close()

        df = pd.concat(li, ignore_index=True)
        df['年級_科目'] = df['年級'] + '_' + df['科目代號']
        df['考試日期'] = test_date
        df['年級'] = df['年級'].astype(int)
        df['班級'] = df['班級'].astype(str)
        df['座號'] = df['座號'].astype(str)

        if db_check:
            st.spinner('Process')
            grade_list = list(df['年級'].unique())
            for a_grade in grade_list:
                st.write(a_grade)
                if a_grade == 1:
                    base_name = st.secrets["base_id_7"]
                elif a_grade == 2:
                    base_name = st.secrets["base_id_8"]
                elif a_grade == 3:
                    base_name = st.secrets["base_id_9"]
                else:
                    st.error('無法判斷年級')
                at.delete_data(base_name)
                temp_df = df[df['年級'] == a_grade]
                at.batch_write_data(temp_df, base_name)

        # for airtable testing purpose
        # i_list = list(df['年級_科目'].unique())
        # for a_sub in i_list:
        #     df_temp = df[df['年級_科目'] == a_sub]
        #     df_temp.to_csv(a_sub + '.csv')

        # Split the 'original_column' into individual characters
        df_split = df['答題對錯'].apply(lambda x: pd.Series(list(x)))

        # Rename the columns with the specified naming convention
        df_split.columns = [f'Q_{str(n).zfill(2)}' for n in range(1, len(df_split.columns) + 1)]

        # Concatenate the split columns with the original DataFrame
        df = pd.concat([df, df_split], axis=1)

        # Drop the original column
        df = df.drop(columns=['答題對錯'])
        # Extract columns starting with 'Q' as value_vars
        value_vars = [col for col in df.columns if col.startswith('Q')]

        # Melt the DataFrame
        melted_df = pd.melt(df, id_vars=[col for col in df.columns if col not in value_vars],
                            value_vars=value_vars, var_name='Question', value_name='Answer')
        melted_df = melted_df.drop('答題狀況', axis=1)
        # Find unique values in column 'AA'
        unique_values = melted_df['年級_科目'].unique()

        # Create a dictionary with 'AA' values as keys and filtered DataFrames as values
        result_dict = {}

        for value in unique_values:
            filtered_df = melted_df[melted_df['年級_科目'] == value].copy()
            st.write(f'Processing {value}')
            result_dict[value] = SharedObjects.Subject(filtered_df)

        st.session_state.subjects = result_dict

        class_dict = {}
        classes = list(melted_df['班級'].unique())

        for a_class in classes:
            class_dict[a_class] = SharedObjects.Class(melted_df[melted_df['班級'] == a_class].copy())
            st.write(f'Processing 班級 {a_class}')

        st.session_state['class_groups'] = class_dict
        st.success('Excel 檔案已經暫存到記憶體，準備進行下一步分析')


if __name__ == '__main__':
    main()
