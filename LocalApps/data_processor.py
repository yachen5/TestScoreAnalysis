# data_processor.py - Backend processing logic
import pandas as pd
from typing import Dict, Tuple, List
from dataclasses import dataclass
from openpyxl import load_workbook

@dataclass
class ProcessedData:
    subjects_dict: Dict
    year_groups: Dict
    class_groups: Dict
    error_ranks: pd.DataFrame

class ExcelDataProcessor:
    def __init__(self, test_date: str):
        self.test_date = test_date

    def process_worksheet(self, excel_file, sheet_name: str) -> pd.DataFrame:
        """Process individual worksheet data"""
        df = pd.read_excel(excel_file, sheet_name=sheet_name, 
                          header=None, engine='openpyxl')
        df = df.dropna(axis=1)
        
        # Set headers and process data
        df.columns = df.iloc[0]
        df = df[1:]
        df['年級'] = df['班級'].astype(str).str[0]
        df['學號'] = 'S' + df['學號'].astype(str)
        
        return df

    def prepare_base_dataframe(self, excel_file) -> pd.DataFrame:
        """Prepare the initial combined DataFrame"""
        wb = load_workbook(excel_file, read_only=True)
        dataframes = []
        
        for sheet_name in wb.sheetnames:
            df = self.process_worksheet(excel_file, sheet_name)
            dataframes.append(df)
        
        wb.close()
        
        # Combine and process DataFrame
        df = pd.concat(dataframes, ignore_index=True)
        df['年級_科目'] = df['年級'] + '_' + df['科目代號']
        df['考試日期'] = self.test_date
        
        # Convert column types
        df['年級'] = df['年級'].astype(int)
        df['班級'] = df['班級'].astype(str)
        df['座號'] = df['座號'].astype(str)
        
        return df

    def process_answer_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Process answer data and create question-answer DataFrame"""
        # Split answers into columns
        df_split = df['答題對錯'].apply(lambda x: pd.Series(list(x)))
        df_split.columns = [f'Q_{str(n).zfill(2)}' for n in range(1, len(df_split.columns) + 1)]
        
        # Combine and melt DataFrame
        df = pd.concat([df, df_split], axis=1)
        df = df.drop(columns=['答題對錯'])
        
        value_vars = [col for col in df.columns if col.startswith('Q')]
        id_vars = [col for col in df.columns if col not in value_vars]
        
        melted_df = pd.melt(df, 
                           id_vars=id_vars,
                           value_vars=value_vars, 
                           var_name='Question', 
                           value_name='Answer')
        
        return melted_df.drop('答題狀況', axis=1)

    def process_excel_file(self, excel_file) -> ProcessedData:
        """Main processing function"""
        # Process base DataFrame
        df = self.prepare_base_dataframe(excel_file)
        melted_df = self.process_answer_data(df)
        
        # Process subjects
        subjects_dict = {}
        error_ranks = []
        
        for subject in melted_df['年級_科目'].unique():
            filtered_df = melted_df[melted_df['年級_科目'] == subject].copy()
            filtered_df = filtered_df.dropna(subset='Answer')
            subject_obj = SharedObjects.Subject(filtered_df, subject)
            subjects_dict[subject] = subject_obj
            error_ranks.append(subject_obj.error_rank)
        
        # Process year groups
        year_groups = {}
        for year in df['年級'].unique():
            year_df = melted_df[melted_df['年級'] == year].copy()
            year_obj = SharedObjects.ClassGroup(year_df, year)
            year_obj.add_pw_by_subject(subjects_dict)
            year_groups[year] = year_obj
        
        # Process class groups
        class_groups = {}
        for class_name in melted_df['班級'].unique():
            class_df = melted_df[melted_df['班級'] == class_name].copy()
            class_groups[class_name] = SharedObjects.Class(class_df)
        
        error_ranks_df = pd.concat(error_ranks, ignore_index=True)
        
        return ProcessedData(
            subjects_dict=subjects_dict,
            year_groups=year_groups,
            class_groups=class_groups,
            error_ranks=error_ranks_df
        )