
import pandas as pd
from openpyxl import load_workbook
from src.models import subject, class_group, student_class
from src.utils import config

def validate_schema(df):
    """Validates that the dataframe contains necessary columns."""
    required_columns = [
        config.COL_CLASS, 
        config.COL_STUDENT_ID, 
        config.COL_SUBJECT_CODE, 
        config.COL_ANSWERS_CORRECTNESS,
        config.COL_SEAT_NUMBER
    ]
    missing = [col for col in required_columns if col not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")
    return True

def process_excel_file(excel_file_path, test_date):
    """
    Processes the uploaded Excel file and prepares data for analysis.

    Args:
        excel_file_path: Path to the uploaded Excel file.
        test_date: The test date provided by the user.

    Returns:
        A dictionary of processed data objects for subjects, years, and classes.
    """
    try:
        wb = load_workbook(excel_file_path, read_only=True)
        worksheet_names = wb.sheetnames
        li = []
    
        for sheet_name in worksheet_names:
            df_t = pd.read_excel(excel_file_path, sheet_name=sheet_name, header=None, engine='openpyxl')
            if df_t.empty:
                continue
                
            df_t = df_t.dropna(axis=1, how='all') # Only drop columns that are ALL NaN
            
            # Use first row as header and remove it
            df_t.columns = df_t.iloc[0]
            df_t = df_t[1:]
            
            # Basic validation before processing
            try:
                validate_schema(df_t)
            except ValueError as e:
                # Add sheet name context to error
                raise ValueError(f"Sheet '{sheet_name}': {str(e)}")
            
            # Standardize columns using config
            if config.COL_CLASS in df_t.columns:
                df_t[config.COL_GRADE] = df_t[config.COL_CLASS].astype(str).str[0]
            
            if config.COL_STUDENT_ID in df_t.columns:
                # Check if 'S' prefix exists, if so do not add again
                df_t[config.COL_STUDENT_ID] = df_t[config.COL_STUDENT_ID].astype(str).apply(
                    lambda x: x if x.startswith(config.PREFIX_STUDENT_ID) else config.PREFIX_STUDENT_ID + x
                )
                
            li.append(df_t)
    
        wb.close()
        
        if not li:
             raise ValueError("No valid data found in the Excel file.")
             
        df = pd.concat(li, ignore_index=True)
        
        # Create Year_Subject column
        df[config.COL_GRADE_SUBJECT] = df[config.COL_GRADE] + '_' + df[config.COL_SUBJECT_CODE]
        df[config.COL_TEST_DATE] = test_date
        
        # Safe conversions
        df[config.COL_GRADE] = pd.to_numeric(df[config.COL_GRADE], errors='coerce').fillna(0).astype(int)
        df[config.COL_CLASS] = df[config.COL_CLASS].astype(str)
        df[config.COL_SEAT_NUMBER] = df[config.COL_SEAT_NUMBER].astype(str)
    
        # Split the '答題對錯' column into individual characters (Q_01, Q_02...)
        # Logic: string of O/X/. chars
        df_split = df[config.COL_ANSWERS_CORRECTNESS].astype(str).apply(lambda x: pd.Series(list(x)))
        
        # Rename Q columns
        df_split.columns = [f'{config.PREFIX_QUESTION}{str(n).zfill(2)}' for n in range(1, len(df_split.columns) + 1)]
    
        # Concatenate and drop original correctness column
        df = pd.concat([df, df_split], axis=1).drop(columns=[config.COL_ANSWERS_CORRECTNESS])
    
        # Melt to long format
        value_vars = [col for col in df.columns if col.startswith(config.PREFIX_QUESTION)]
        melted_df = pd.melt(df, id_vars=[col for col in df.columns if col not in value_vars],
                            value_vars=value_vars, var_name=config.COL_QUESTION, value_name=config.COL_ANSWER)
        
        if config.COL_ANSWER_STATUS in melted_df.columns:
            melted_df = melted_df.drop(config.COL_ANSWER_STATUS, axis=1)
    
        # Create Subject objects
        result_dict = {}
        unique_values = melted_df[config.COL_GRADE_SUBJECT].unique()
        for value in unique_values:
            filtered_df = melted_df[melted_df[config.COL_GRADE_SUBJECT] == value].copy()
            filtered_df = filtered_df.dropna(subset=config.COL_ANSWER)
            result_dict[value] = subject.Subject(filtered_df, value)
    
        # Create Year Group objects
        year_dict = {}
        for a_year in list(df[config.COL_GRADE].unique()):
            year_obj = class_group.ClassGroup(melted_df[melted_df[config.COL_GRADE] == a_year].copy(), a_year)
            year_obj.add_pw_by_subject(result_dict)
            year_dict[a_year] = year_obj
    
        # Create Class objects
        class_dict = {}
        for a_class in list(melted_df[config.COL_CLASS].unique()):
            class_dict[a_class] = student_class.Class(melted_df[melted_df[config.COL_CLASS] == a_class].copy())
    
        return {
            'subjects': result_dict,
            'year_groups': year_dict,
            'class_groups': class_dict
        }
    except Exception as e:
        # Re-raise with meaningful context if it's not already a ValueError we raised
        if isinstance(e, ValueError):
            raise e
        raise RuntimeError(f"Error processing Excel file: {str(e)}") from e

