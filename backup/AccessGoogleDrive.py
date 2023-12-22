import pandas as pd
import streamlit as st
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from openpyxl import load_workbook

# If modifying these scopes, delete the file token.json.
SCOPES = ['https://www.googleapis.com/auth/drive']


def authenticate():
    creds = None
    # The file token.json stores the user's access and refresh tokens.
    try:
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    except FileNotFoundError:
        flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
        creds = flow.run_local_server(port=0)
        with open('../token.json', 'w') as token:
            token.write(creds.to_json())
    return creds


def download_file(file_id, filename):
    creds = authenticate()
    service = build('drive', 'v3', credentials=creds)

    request = service.files().get_media(fileId=file_id)

    with open(filename, 'wb') as file:
        downloader = request.execute()
        file.write(downloader)


def list_files():
    creds = authenticate()
    service = build('drive', 'v3', credentials=creds)

    # Replace 'Folder Name' with your folder name or ID.
    folder_id = 'to check google'

    # Retrieve the files in the folder
    results = service.files().list(q=f"'{folder_id}' in parents",
                                   pageSize=10, fields="nextPageToken, files(id, name)").execute()
    items = results.get('files', [])

    xlsx_list = []
    if not items:
        st.write('No files found.')
    else:
        st.write('Files:')
        for item in items:
            if item['name'].endswith('.xlsx'):
                xlsx_list.append(item)
    a_file = st.selectbox('Files', xlsx_list)

    download_file(a_file['id'], f'{a_file["name"]}')
    excel_file_path = a_file['name']
    # df = pd.read_excel(a_file['name'], engine='openpyxl')

    wb = load_workbook(a_file['name'], read_only=True)

    # Get the names of all worksheets in the Excel file
    worksheet_names = wb.sheetnames

    # Initialize an empty DataFrame to store the combined data
    li = []

    # Loop through each worksheet and append its data to the combined DataFrame
    for sheet_name in worksheet_names:
        # Read data from the current worksheet
        df_t = pd.read_excel(excel_file_path, sheet_name=sheet_name, header=None, engine='openpyxl')
        df_t = df_t.dropna(axis=1)

        # Use the first row as header
        df_t.columns = df_t.iloc[0]

        # Drop the first row (since it's now the header)
        df_t = df_t[1:]
        df_t['年級'] = df_t['班級'].astype(str).str[0]
        df_t['學號'] = 'S' + df_t['學號'].astype(str)

        # Append the data to the combined DataFrame
        li.append(df_t)

    # Close the workbook
    wb.close()

    df = pd.concat(li, ignore_index=True)
    st.dataframe(df)


if __name__ == '__main__':
    list_files()
